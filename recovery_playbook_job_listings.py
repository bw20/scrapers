import requests, bs4, time, openpyxl, re
from random import randint
from time import sleep
from openpyxl.utils import get_column_letter
from datetime import datetime
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains as ac
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def job_listings():
    #PROXY = "81.200.63.108:60579"
    #webdriver.DesiredCapabilities.FIREFOX['proxy'] = {
        #"httpProxy": PROXY,
        #"ftpProxy": PROXY,
        #"sslProxy": PROXY,
        #"proxyType": "MANUAL",
    #}
    browser = webdriver.Firefox()
    url = 'https://www.seek.com.au/jobs'
    browser.get(url)
    list_of_jobs = ['Accounting','Administration & Office Support','Advertising, Arts & Media','Banking & Financial Services','Call Centre & Customer Service','CEO & General Management','Community Services & Development','Construction','Consulting & Strategy','Design & Architecture','Education & Training','Engineering','Farming, Animals & Conservation','Government & Defence','Healthcare & Medical','Hospitality & Tourism','Human Resources & Recruitment','Information & Communication Technology','Insurance & Superannuation','Legal','Manufacturing, Transport & Logistics','Marketing & Communications','Mining, Resources & Energy','Real Estate & Property','Retail & Consumer Products','Sales','Science & Technology','Self Employment','Sport & Recreation','Trades & Service']
    list_of_states = ['National', 'ACT', 'New South Wales', 'Northern Territory', 'Queensland', 'South Australia', 'Tasmania', 'Victoria', 'Western Australia']
    states = {}
    state_search_count = 0
    for state in list_of_states:
        data = {} #this needs to be reset every time the loop starts again.
        #time.sleep(randint(3, 5))
        blank_click = browser.find_element_by_xpath('//html')
        blank_click.click()
        job_classification = browser.find_element_by_css_selector('._3_QSUXU') #_1tb9M9B
        job_classification.click()
        bs4_browser = bs4.BeautifulSoup(browser.page_source, 'html.parser')
        #This bit scrapes the total job figure from the page
        job_listings = bs4_browser('strong', '_7ZnNccT')[2].getText()#lwHBT6d
        job_label = 'Any Classification'
        data[job_label] = int(job_listings.replace(',', '')) #this will be {'All classifications' : '1,234'}
        for i in range(1, len(list_of_jobs) + 1):
            #This for loop goes through each list number in the dropdown and records them
            job_label = bs4_browser.select('li._3i3Jw55:nth-child(' + str(i) + ') > a:nth-child(1) > span:nth-child(1)')[0].getText()
            job_listings = bs4_browser.select('li._3i3Jw55:nth-child('+ str(i) +') > span:nth-child(2)')[0].getText()
            data[job_label] = int(job_listings[1:].replace(',', ''))
            time.sleep(randint(1, 2)) #this can slow down the scrape quite a lot
        states[state] = data #This should be: {'National: {'All classifications : 1,234, ACT : 1,234 etc.}}
        #this bit sets the state that we are scraping, it types the state and hits enter. It's here because the first run-through should scrape the national figures
        print(state, 'has finished scraping.')
        print(states)
        time.sleep(randint(5, 60))
        job_classification.click()
        state_search = browser.find_element_by_css_selector('#SearchBar__Where')
        if state != 'National':
            delete = browser.find_element_by_css_selector('span._2LBRxrs:nth-child(3) > div:nth-child(1) > span:nth-child(1) > svg:nth-child(1)')
            delete.click()
        state_search_count += 1
        if state_search_count < len(list_of_states):
            state_search.send_keys(list_of_states[state_search_count])
            state_search.send_keys(Keys.RETURN)
    print('All cases finished scraping.')
    return states
def write_to_excel(data, week_number):
    wb = openpyxl.load_workbook('recovery_playbook.xlsx')
    sheet = wb['seek']
    week = 'Week ' + str(week_number)
    count = 0
    states = {'National': 'National', 'ACT': 'ACT', 'New South Wales': 'NSW', 'Northern Territory': 'NT', 'Queensland': 'QLD', 'South Australia': 'SA', 'Tasmania': 'TAS', 'Victoria': 'VIC', 'Western Australia': 'WA'}
    for location in data:
        for job in data[location]:
            #print(job, data[location][job], states[location], week)
            row_number = sheet.max_row + 1
            sheet['C' + str(row_number)] = job
            sheet['D' + str(row_number)] = data[location][job]
            sheet['E' + str(row_number)] = states[location]
            sheet['F' + str(row_number)] = week
    wb.save(filename = str(date.today()) + '_recovery_playbook.xlsx')
    file = str(date.today()) + '_recovery_playbook.xlsx'
    print('Saved.')
    return file
def email_results(subject, send_to, attach_file, message):
    import email, smtplib, ssl
    from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.utils import formatdate
    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From'] = 'bayee2006@hotmail.com'        #email_user
    msg['To'] = send_to
    msg['Date'] = formatdate(localtime = True)
    msg.attach(MIMEText(message))
    part = MIMEBase('application', 'octet-stream')
    filename = attach_file
    part.set_payload(open(filename, 'rb').read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename=attach_file)
    msg.attach(part)
    s = smtplib.SMTP('smtp-mail.outlook.com', 587)
    s.ehlo()
    s.starttls()
    pw = str(input('Password: '))
    s.login('bayee2006@hotmail.com', pw) #change the PW here
    s.sendmail('bayee2006@hotmail.com', 'bayee2006@hotmail.com', msg.as_string())
    s.quit()
    print('mail sent. check your inbox')

#data = job_listings()
data = {'National': {'Any Classification': 167134, 'Accounting': 6007, 'Administration & Office Support': 8077, 'Advertising, Arts & Media': 600, 'Banking & Financial Services': 3161, 'Call Centre & Customer Service': 2917, 'CEO & General Management': 462, 'Community Services & Development': 6961, 'Construction': 6276, 'Consulting & Strategy': 688, 'Design & Architecture': 1467, 'Education & Training': 7278, 'Engineering': 4144, 'Farming, Animals & Conservation': 1249, 'Government & Defence': 5392, 'Healthcare & Medical': 17661, 'Hospitality & Tourism': 11257, 'Human Resources & Recruitment': 3447, 'Information & Communication Technology': 13101, 'Insurance & Superannuation': 800, 'Legal': 2407, 'Manufacturing, Transport & Logistics': 13105, 'Marketing & Communications': 3402, 'Mining, Resources & Energy': 4750, 'Real Estate & Property': 2958, 'Retail & Consumer Products': 6560, 'Sales': 7225, 'Science & Technology': 754, 'Self Employment': 48, 'Sport & Recreation': 1097, 'Trades & Services': 23574}, 'ACT': {'Any Classification': 5362, 'Accounting': 6007, 'Administration & Office Support': 8077, 'Advertising, Arts & Media': 600, 'Banking & Financial Services': 3161, 'Call Centre & Customer Service': 2917, 'CEO & General Management': 462, 'Community Services & Development': 6961, 'Construction': 6276, 'Consulting & Strategy': 688, 'Design & Architecture': 1467, 'Education & Training': 7278, 'Engineering': 4144, 'Farming, Animals & Conservation': 1249, 'Government & Defence': 5392, 'Healthcare & Medical': 17661, 'Hospitality & Tourism': 11257, 'Human Resources & Recruitment': 3447, 'Information & Communication Technology': 13101, 'Insurance & Superannuation': 800, 'Legal': 2407, 'Manufacturing, Transport & Logistics': 13105, 'Marketing & Communications': 3402, 'Mining, Resources & Energy': 4750, 'Real Estate & Property': 2958, 'Retail & Consumer Products': 6560, 'Sales': 7225, 'Science & Technology': 754, 'Self Employment': 48, 'Sport & Recreation': 1097, 'Trades & Services': 23574}, 'New South Wales': {'Any Classification': 52363, 'Accounting': 2206, 'Administration & Office Support': 2451, 'Advertising, Arts & Media': 281, 'Banking & Financial Services': 1513, 'Call Centre & Customer Service': 1024, 'CEO & General Management': 137, 'Community Services & Development': 2336, 'Construction': 1799, 'Consulting & Strategy': 292, 'Design & Architecture': 578, 'Education & Training': 2183, 'Engineering': 1123, 'Farming, Animals & Conservation': 345, 'Government & Defence': 1604, 'Healthcare & Medical': 5488, 'Hospitality & Tourism': 3619, 'Human Resources & Recruitment': 1270, 'Information & Communication Technology': 5063, 'Insurance & Superannuation': 371, 'Legal': 931, 'Manufacturing, Transport & Logistics': 3635, 'Marketing & Communications': 1396, 'Mining, Resources & Energy': 446, 'Real Estate & Property': 997, 'Retail & Consumer Products': 1949, 'Sales': 2562, 'Science & Technology': 234, 'Self Employment': 24, 'Sport & Recreation': 341, 'Trades & Services': 6165}, 'Northern Territory': {'Any Classification': 1882, 'Accounting': 28, 'Administration & Office Support': 81, 'Advertising, Arts & Media': 13, 'Banking & Financial Services': 10, 'Call Centre & Customer Service': 12, 'CEO & General Management': 9, 'Community Services & Development': 133, 'Construction': 86, 'Consulting & Strategy': 6, 'Design & Architecture': 4, 'Education & Training': 122, 'Engineering': 48, 'Farming, Animals & Conservation': 24, 'Government & Defence': 104, 'Healthcare & Medical': 235, 'Hospitality & Tourism': 165, 'Human Resources & Recruitment': 26, 'Information & Communication Technology': 69, 'Insurance & Superannuation': 4, 'Legal': 12, 'Manufacturing, Transport & Logistics': 104, 'Marketing & Communications': 6, 'Mining, Resources & Energy': 132, 'Real Estate & Property': 18, 'Retail & Consumer Products': 73, 'Sales': 32, 'Science & Technology': 3, 'Self Employment': 2, 'Sport & Recreation': 10, 'Trades & Services': 311}, 'Queensland': {'Any Classification': 34135, 'Accounting': 1137, 'Administration & Office Support': 1697, 'Advertising, Arts & Media': 99, 'Banking & Financial Services': 511, 'Call Centre & Customer Service': 581, 'CEO & General Management': 76, 'Community Services & Development': 1510, 'Construction': 1353, 'Consulting & Strategy': 74, 'Design & Architecture': 244, 'Education & Training': 1743, 'Engineering': 922, 'Farming, Animals & Conservation': 298, 'Government & Defence': 914, 'Healthcare & Medical': 3716, 'Hospitality & Tourism': 2385, 'Human Resources & Recruitment': 596, 'Information & Communication Technology': 1933, 'Insurance & Superannuation': 115, 'Legal': 506, 'Manufacturing, Transport & Logistics': 2846, 'Marketing & Communications': 537, 'Mining, Resources & Energy': 862, 'Real Estate & Property': 633, 'Retail & Consumer Products': 1287, 'Sales': 1472, 'Science & Technology': 142, 'Self Employment': 12, 'Sport & Recreation': 219, 'Trades & Services': 5715}, 'South Australia': {'Any Classification': 8766, 'Accounting': 269, 'Administration & Office Support': 486, 'Advertising, Arts & Media': 15, 'Banking & Financial Services': 123, 'Call Centre & Customer Service': 171, 'CEO & General Management': 39, 'Community Services & Development': 414, 'Construction': 369, 'Consulting & Strategy': 25, 'Design & Architecture': 41, 'Education & Training': 279, 'Engineering': 288, 'Farming, Animals & Conservation': 116, 'Government & Defence': 347, 'Healthcare & Medical': 948, 'Hospitality & Tourism': 626, 'Human Resources & Recruitment': 166, 'Information & Communication Technology': 417, 'Insurance & Superannuation': 42, 'Legal': 79, 'Manufacturing, Transport & Logistics': 804, 'Marketing & Communications': 111, 'Mining, Resources & Energy': 225, 'Real Estate & Property': 108, 'Retail & Consumer Products': 375, 'Sales': 326, 'Science & Technology': 41, 'Self Employment': 4, 'Sport & Recreation': 48, 'Trades & Services': 1464},'Tasmania': {'Any Classification': 1922, 'Accounting': 29, 'Administration & Office Support': 80, 'Advertising, Arts & Media': 9, 'Banking & Financial Services': 20, 'Call Centre & Customer Service': 36, 'CEO & General Management': 11, 'Community Services & Development': 94, 'Construction': 54, 'Consulting & Strategy': 7, 'Design & Architecture': 8, 'Education & Training': 86, 'Engineering': 46, 'Farming, Animals & Conservation': 38, 'Government & Defence': 120, 'Healthcare & Medical': 291, 'Hospitality & Tourism': 174, 'Human Resources & Recruitment': 21, 'Information & Communication Technology': 40, 'Insurance & Superannuation': 5, 'Legal': 17, 'Manufacturing, Transport & Logistics': 144, 'Marketing & Communications': 24, 'Mining, Resources & Energy': 54, 'Real Estate & Property': 22, 'Retail & Consumer Products': 100, 'Sales': 70, 'Science & Technology': 8, 'Self Employment': 0, 'Sport & Recreation': 7, 'Trades & Services': 307}, 'Victoria': {'Any Classification': 42670, 'Accounting': 1608, 'Administration & Office Support': 2136, 'Advertising, Arts & Media': 133, 'Banking & Financial Services': 766, 'Call Centre & Customer Service': 863, 'CEO & General Management': 124, 'Community Services & Development': 1659, 'Construction': 1489, 'Consulting & Strategy': 176, 'Design & Architecture': 461, 'Education & Training': 1940, 'Engineering': 969, 'Farming, Animals & Conservation': 268, 'Government & Defence': 1254, 'Healthcare & Medical': 4938, 'Hospitality & Tourism': 2953, 'Human Resources & Recruitment': 869, 'Information & Communication Technology': 3362, 'Insurance & Superannuation': 202, 'Legal': 579, 'Manufacturing, Transport & Logistics': 3867, 'Marketing & Communications': 1071, 'Mining, Resources & Energy': 183, 'Real Estate & Property': 880, 'Retail & Consumer Products': 1805, 'Sales': 1981, 'Science & Technology': 212, 'Self Employment': 4, 'Sport & Recreation': 324, 'Trades & Services': 5594}, 'Western Australia': {'Any Classification': 19526, 'Accounting': 573, 'Administration & Office Support': 932, 'Advertising, Arts & Media': 38, 'Banking & Financial Services': 163, 'Call Centre & Customer Service': 192, 'CEO & General Management': 45, 'Community Services & Development': 683, 'Construction': 980, 'Consulting & Strategy': 65, 'Design & Architecture': 91, 'Education & Training': 646, 'Engineering': 638, 'Farming, Animals & Conservation': 151, 'Government & Defence': 415, 'Healthcare & Medical': 1605, 'Hospitality & Tourism': 988, 'Human Resources & Recruitment': 408, 'Information & Communication Technology': 762, 'Insurance & Superannuation': 46, 'Legal': 204, 'Manufacturing, Transport & Logistics': 1564, 'Marketing & Communications': 172, 'Mining, Resources & Energy': 2764, 'Real Estate & Property': 219, 'Retail & Consumer Products': 789, 'Sales': 625, 'Science & Technology': 80, 'Self Employment': 2, 'Sport & Recreation': 121, 'Trades & Services': 3565}}

write_to_excel(data, 39)
#email_results('Job listings results', 'bayee2006@hotmail.com', file, 'See attached.')